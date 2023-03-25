# todo:
# + dữ liệu cần nhập: độ dài tấm chắn sáng
# -- va chạm mềm: t1, t2', m1, m2, quảng đường
# - tự hiểu v2 = 0
# -- va chạm đàn hồi: m1, m2, t1', t2', quãng đường

# =====

# thu vien gui
import PySimpleGUI as sg
# thu vien lay du lieu serial
import math
import os
import datetime
import serial
import serial.tools.list_ports
# thu vien chinh sua file excel
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl import Workbook
# thu vien hinh anh
from PIL import Image, ImageTk

# TODO: ham xuat du lieu thi nghiem ra file excel
def export_to_excel(inelastic_data, inelastic_headings, elastic_data, elastic_headings, path):
    # create a workbook
    wb = Workbook()
    ws = wb.active
    
    # main data
    rows = []
    rows.append(["Va chạm mềm"])
    rows.append(inelastic_headings)
    for d in inelastic_data: rows.append(d)

    rows.append([])    

    rows.append(["Va chạm đàn hồi"])
    rows.append(elastic_headings)
    for d in elastic_data: rows.append(d)

    for r in rows: ws.append(r)

    # inelsatic_table
    cname1 = ws.cell(1, 1)
    cname1.alignment = Alignment(horizontal='center')
    cname1.fill = PatternFill(start_color='ffeb9c', end_color='ffeb9c', fill_type='solid')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(inelastic_headings))
    for i in range(len(inelastic_headings)):
        ws.cell(row=2, column=i+1).fill = PatternFill(start_color="bdd7ee", end_color="bdd7ee", fill_type='solid')
    for i in range(2, len(inelastic_data)+2):
        for j in range(len(inelastic_headings)):
            ws.cell(row=i+1, column=j+1).fill = PatternFill(start_color="fff2cc", end_color="fff2cc", fill_type='solid')
    # elatstic_table
    cname2 = ws.cell(row=len(inelastic_data)+4, column=1)
    cname2.alignment = Alignment(horizontal='center')
    cname2.fill = PatternFill(start_color='c6efce', end_color='c6efce', fill_type='solid')
    ws.merge_cells(start_row=len(inelastic_data)+4, start_column=1, end_row=len(inelastic_data)+4, end_column=len(elastic_headings))
    for i in range(len(inelastic_headings)):
        ws.cell(row=len(inelastic_data)+5, column=i+1).fill = PatternFill(start_color="bdd7ee", end_color="bdd7ee", fill_type='solid')
    for i in range(len(inelastic_data)+5, len(elastic_data)+len(inelastic_data)+5):
        for j in range(len(elastic_headings)):
            ws.cell(row=i+1, column=j+1).fill = PatternFill(start_color="fff2cc", end_color="fff2cc", fill_type='solid')

    wb.save(path)

    return 0

# trinh chon cong com
def portselector():
    # hien thi tat ca cong serial dang mo tren may tinh
    ports = serial.tools.list_ports.comports()
    ports = sorted(ports)

    portlist = []
    for i in range(len(ports)):
        port = ports[i].name
        desc = ports[i].description
        hwid = ports[i].hwid
        portlist.append("{}. {}: {} [{}] \n".format(i+1, port, desc, hwid))
        # print("{}: {} [{}]".format(port, desc, hwid))

    # print(portlist)

    layout = [
        [sg.Text("Chọn cổng COM đến ESP32: ", background_color='#eeeeee', text_color='#000')],
        [sg.Combo(values=portlist, expand_x=True, background_color='#eeeeee', text_color='#000', button_background_color='#eeeeee', button_arrow_color="#000")],
        [sg.Submit(button_text="Kết nối", button_color=('#fff', '#000'))]
    ]
    win = sg.Window("Chọn cổng COM", layout, finalize=True, background_color='#eeeeee', font=("Arial", 10))
    e, v = win.read()
    win.close()

    # chon cong com den esp32
    # i = int(input("Chọn cổng COM để kết nối đến ESP32: "))
    port = ports[portlist.index(v[0])]
    ser = serial.Serial(str(port.name), 9600, timeout=0.050)
    return ser, str(port.description)


# mo cua so portselector
ser, ser_desc = portselector()

# data

# bien du trang thai may do
isCounting = False

# bien giu ten tab dang chon (bang so lieu nao)
tg_tab = ""

# du lieu cua thi nghiem va cham mem
inelastic_tries = 1
# test data
data_inelastic = [[1, 0.21, 0.21, 0.179, 0.559, 0.117, 0, 0.117, 0.354, 0.282, 0.118], [2, 0.21, 0.31, 0.179, 0.559, 0.117, 0, 0.117, 0.354, 0.282, 0.147], [3, 0.21, 0.41, 0.179, 0.559, 0.117, 0, 0.117, 0.354, 0.282, 0.175]]
# data_inelastic = []
headings_inelastic = ["Lần", "m1", "m2", "t1", "v1", "p1", "p2", "p", "t2'", "v1'=v2'", "p'"]

# du lieu cua thi nghiem va cham dan hoi
elastic_tries = 1
# test data
data_elastic = [[1, 0.21, 0.21, 0, 0, 0.179, 0.354, 0.559, 0.282, 0.117, -0.059], [2, 0.21, 0.31, 0, 0, 0.179, 0.354, 0.559, 0.282, 0.117, -0.087], [3, 0.21, 0.41, 0, 0, 0.179, 0.354, 0.559, 0.282, 0.117, -0.116]]
# data_elastic = []
headings_elastic = ["Lần", "m1", "m2", "p1", "p2", "t1'", "t2'", "v1'", "v2'", "p1'", "p'"]

# ham su li thong tin dua vao
# va cham mem
def datain_inelastic(tries, t1, t2_s, m1, m2, s):
    v1 = round(s/t1, 3)
    p1 = round(v1*m1, 3)
    p2 = 0
    p = p1+p2
    v_s = round(s/t2_s, 3)
    p_s = round(v_s*(m1+m2), 3)
    return [tries, m1, m2, t1, v1, p1, p2, p, t2_s, v_s, p_s]

# va cham dan hoi
def datain_elastic(tries, t1_s, t2_s, m1, m2, s):
    p1 = p2 = 0
    v1_s = round(s/t1_s, 3)
    v2_s = round(s/t2_s, 3)
    p1_s = round(v1_s*m1, 3)
    p2_s = -1*round(v2_s*m2, 3)
    return [tries, m1, m2, p1, p2, t1_s, t2_s, v1_s, v2_s, p1_s, p2_s]

# ham su li thong tin chinh
def datain(exp_mode, tries):
    sout = ser.readline()
    sout_decoded = str(sout).split(";")
    # 0; in1; in2; mode
    print(sout_decoded)
    if (len(sout_decoded) < 2): 
        return "-1"
    else:
        in1, in2 = float(int(sout_decoded[1])/1000), float(int(sout_decoded[2])/1000)
        while True:
            # tạo input box trống -> đặt key -> update nội dung theo key
            # dùng hàm ngoài để parse công thức
            layout = [
                [sg.Text(f"Nhập dữ liệu còn thiếu của lần đo thứ {tries}:",  background_color='#eeeeee', text_color='#000')],
                [sg.Text(f"Dữ liệu thời gian từ bộ đo: [{in1}], [{in2}] (ms)",  background_color='#eeeeee', text_color='#000')],
                [sg.Text("m1:                   ",  background_color='#eeeeee', text_color='#000'), sg.InputText( background_color='#fff', text_color='#000', border_width=0)],
                [sg.Text("m2:                   ",  background_color='#eeeeee', text_color='#000'), sg.InputText( background_color='#fff', text_color='#000', border_width=0)],
                [sg.Text("Quãng đường: ",  background_color='#eeeeee', text_color='#000'), sg.InputText( background_color='#fff', text_color='#000', border_width=0)],
                [sg.Submit(button_text="Hoàn tất",  button_color=('#fff', '#000'), bind_return_key=True)]
            ]
            win = sg.Window("Nhập dữ liệu đo", layout, finalize=True, background_color='#eeeeee', font=('Arial', 14), keep_on_top=True)
            e, v = win.read()
            win.close()
            if (v[0] != "" and v[1] != "" and v[2] != ""): 
                break
            else: sg.Popup("Các ô dữ liệu không được để trống!", title="Chú ý", background_color='#eeeeee', text_color='#000', button_color=('#fff', '#000'))

        # m1, m2, s tu nhap
        m1, m2, s = float(v[0]), float(v[1]), float(v[2])
        if exp_mode == "elastic":
            data_elastic.append(datain_elastic(tries, in1, in2, m1, m2, s))
        elif exp_mode == "inelastic":
            data_inelastic.append(datain_inelastic(tries, in1, in2, m1, m2, s))
        

# gui
# define menu selections
menu = [
        ['&Tệp', ['&Xuất đồ thị...', '&Thoát']],
        ['&Số liệu', ['&Bảng số liệu', ['Xóa bảng', 'Xóa dòng', 'Bật chỉnh sửa']]],
        ['&Trợ giúp', ['&Thông tin']]
    ]

# define table layout
tab_table_inelasitc = [
    # bang so lieu
    # thêm khung/màu so le
    [sg.Table(
        values=data_inelastic, 
        headings=headings_inelastic, 
        key="-t_inelastic-", 
        auto_size_columns=False, 
        num_rows=10, 
        justification="center", 
        expand_x=True, 
        expand_y=True, 
        font=("Arial", 14, "bold"), 
        # header_background_color=(), header_text_color=(),
        alternating_row_color = "#add8e6",
        selected_row_colors = ("#000", "#86a8b3"),
        header_relief=sg.RELIEF_SOLID,
        background_color='#fff', 
        text_color='#000',
        sbar_trough_color='#fff', 
        sbar_background_color='#eeeeee', 
        sbar_arrow_color='#fff', 
        sbar_frame_color='#eeeeee', 
        sbar_relief=sg.RELIEF_FLAT,
        enable_events=True
    )],
]

tab_table_elasitc = [
    # bang so lieu
    # thêm khung/màu so le
    [sg.Table(
        values=data_elastic, 
        headings=headings_elastic, 
        key="-t_elastic-", 
        auto_size_columns=False,  
        num_rows=10, 
        justification="center", 
        expand_x=True, 
        expand_y=True, 
        font=("Arial", 14, "bold"), 
        # header_background_color=(), header_text_color=(),
        alternating_row_color = "#add8e6",
        selected_row_colors = ("#000", "#86a8b3"),
        header_relief=sg.RELIEF_SOLID,
        background_color='#fff', 
        text_color='#000',
        sbar_trough_color='#fff', 
        sbar_background_color='#eeeeee', 
        sbar_arrow_color='#fff', 
        sbar_frame_color='#eeeeee', 
        sbar_relief=sg.RELIEF_FLAT,
        enable_events=True
    )],
]

# main layout
layout = [
    [sg.Menu(menu, tearoff=False, key='-menu-')],
    [sg.TabGroup([
        [
            sg.Tab('Va chạm mềm', tab_table_inelasitc, background_color='#eeeeee', key='-tab_inelastic-'),
            sg.Tab('Va chạm đàn hồi', tab_table_elasitc, background_color='#eeeeee', key='-tab_elastic-'),
        ]
    ], expand_x=True, expand_y=True, tab_background_color="#eeeeee", background_color="#eeeeee", key="-tg-")],
    [sg.Button("Bắt đầu đo", key="-start-")],
    [sg.StatusBar("Trạng thái đo: Sẵn sàng", key="-status-", text_color="#000", background_color="#eeeeee", relief=sg.RELIEF_FLAT, size=(1, 1), expand_x=True,  justification="left")],
]

# tao cua so chuong trinh chinh
win = sg.Window(f"Kết quả đo - {ser.name}: {ser_desc}", layout, resizable=True, finalize=True, background_color='#eeeeee')
# main loop
while True:
    try:

        e, v = win.read(timeout=250)
        tg_tab = v["-tg-"]
        win["-status-"].update("Trạng thái đo: Sẵn sàng" if not isCounting else "Trạng thái đo: Đang đo")
        if e == sg.WINDOW_CLOSED or e == "Thoát":
            win.close()
            break

        if e == "Xuất đồ thị...":
            dir = str(os.getcwd())
            name = datetime.datetime.now()
            name = name.strftime("%d%m%y_%H%M%S") + ".xlsx"
            layout = [
                [
                    sg.Text("Chọn thư mục: ", background_color='#eeeeee', text_color='#000'), 
                    sg.Input(key="-IN2-" ,change_submits=True, default_text=dir, background_color='#fff', text_color='#000', border_width=0), 
                    sg.FolderBrowse(key="-IN-", button_color=('#fff', '#000'))
                ],
                [
                    sg.Button("Chọn", key="Submit", button_color=('#fff', '#000'))
                ]
            ]
            exp_win = sg.Window("Xuất đồ thị", layout, finalize=True, background_color='#eeeeee')
            while True:
                event, values = exp_win.read()
                if event == sg.WIN_CLOSED or event=="Exit":
                    break
                elif event == "Submit":
                    dir = values["-IN2-"]
                    dir = dir + f"/{name}"
                    export_to_excel(data_inelastic, headings_inelastic, data_elastic, headings_elastic, dir)
                    sg.Popup(f"Đã xuất {name} tại đường dẫn {dir}.", title="Hoàn tất", background_color='#eeeeee', text_color='#000', button_color=('#fff', '#000'))
                    break
            exp_win.close()

        if e == "Xóa bảng":
            data_elastic = data_inelastic = []
            win["-t_inelastic-"].update(values=data_inelastic)
            win["-t_elastic-"].update(values=data_elastic)
        if e == "Xóa dòng":
            cfg_layout = [
                [
                    sg.Text("Dòng: ", background_color='#eeeeee', text_color='#000'), 
                    sg.In(background_color='#fff', text_color='#000', border_width=0)
                ],
                [sg.Button("Hoàn tất", key="-cfg_done-", button_color=('#fff', '#000'), bind_return_key=True)]
            ]
            cfg = sg.Window("Xóa", cfg_layout, element_justification='left', background_color='#eeeeee', font=("Arial", 10), finalize=True)
            cfge, cfgv = cfg.read()
            if cfge == "-cfg_done-":
                if tg_tab == "-tab_inelastic-":
                    data_inelastic.pop(int(cfgv[0])-1)
                    win["-t_inelastic-"].update(values=data_inelastic)
                if tg_tab == "-tab_elastic-":
                    data_elastic.pop(int(cfgv[0])-1)
                    win["-t_elastic-"].update(values=data_elastic)
                # uwu
                cfg.close()
        
        if e == "-start-":
            if not isCounting:
                ser.write("s-ab\n".encode())
                isCounting = not isCounting
            else:
                sg.Popup("Tiến trình đo vẫn đang chạy trên máy đo!", title="Thông báo", background_color='#eeeeee', text_color='#000', button_color=('#fff', '#000'))

        if (ser.in_waiting != 0):
            if tg_tab == "-tab_inelastic-": 
                exp_mode = "inelastic"
                if (datain(exp_mode, inelastic_tries) != "-1"):
                    isCounting = False
                    inelastic_tries += 1 
                    win["-t_inelastic-"].update(values=data_inelastic)
                    print(data_inelastic)
            elif tg_tab == "-tab_elastic-": 
                exp_mode = "elastic"
                if (datain(exp_mode, elastic_tries) != "-1"):
                    isCounting = False
                    elastic_tries += 1 
                    win["-t_elastic-"].update(values=data_elastic)
                    print(data_elastic)


    except serial.serialutil.SerialException:
        sg.Popup("Thiết bị đo đã ngắt kết nối!", title="Thông báo", background_color='#eeeeee', text_color='#000', button_color=('#fff', '#000'))
        break
    
    except Exception as e: 
        print(e)
        break

# dung chuong trinh sau khi thoat phan mem
raise SystemExit(0)